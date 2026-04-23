
import re
import unicodedata
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

PREVIEW_SHEET = "parsed_summary_preview"
COMPARISON_SHEET = "comparison"
UNRECOGNIZED_SHEET = "unrecognized_lines"
ELIGIBLE_SHEET = "eligible_players"
RAW_SHEET = "export_raw"
RULES_SHEET = "rules"

def safe_float(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        try:
            if pd.isna(value):
                return None
        except Exception:
            pass
        return float(value)
    s = str(value).strip()
    if not s:
        return None
    s = s.replace("\xa0", " ").replace("$", "").replace("€", "").replace("₽", "").replace(" ", "")
    if "," in s and "." in s:
        s = s.replace(",", "")
    else:
        s = s.replace(",", ".")
    try:
        return float(s)
    except Exception:
        return None

def safe_int(value):
    f = safe_float(value)
    if f is None:
        return None
    return int(round(f))

def normalize_spaces(s: str) -> str:
    return re.sub(r"\s+", " ", str(s or "")).strip()

def strip_accents(text: str) -> str:
    text = unicodedata.normalize("NFKD", text)
    return "".join(ch for ch in text if not unicodedata.combining(ch))

COUNTRY_ALIASES = {
    "korea (south)": "south korea", "south korea": "south korea", "republic of korea": "south korea",
    "korea, republic of": "south korea", "venezuela (bolivarian republic)": "venezuela",
    "venezuela, bolivarian republic of": "venezuela", "venezuela": "venezuela",
    "russian federation": "russia", "russia": "russia",
    "cote d'ivoire": "cote d'ivoire", "côte d'ivoire": "cote d'ivoire", "ivory coast": "cote d'ivoire",
    "cote divoire": "cote d'ivoire", "kyrgyz republic": "kyrgyzstan", "kyrgyzstan": "kyrgyzstan",
    "lao people's democratic republic": "laos", "laos": "laos", "syrian arab republic": "syria",
    "syria": "syria", "iran (islamic republic of)": "iran", "iran": "iran",
    "moldova, republic of": "moldova", "moldova": "moldova",
    "tanzania, united republic of": "tanzania", "tanzania": "tanzania",
    "bolivia (plurinational state of)": "bolivia", "bolivia": "bolivia",
    "viet nam": "vietnam", "vietnam": "vietnam", "uae": "united arab emirates",
    "united arab emirates": "united arab emirates", "usa": "united states", "us": "united states",
    "united states of america": "united states", "united states": "united states",
    "uk": "united kingdom", "great britain": "united kingdom", "united kingdom": "united kingdom",
}

GEO_CODE_MAP = {
    "AE":"United Arab Emirates","AF":"Afghanistan","AL":"Albania","AM":"Armenia","AO":"Angola","AR":"Argentina",
    "AT":"Austria","AU":"Australia","AZ":"Azerbaijan","BA":"Bosnia and Herzegovina","BD":"Bangladesh","BE":"Belgium",
    "BF":"Burkina Faso","BG":"Bulgaria","BH":"Bahrain","BJ":"Benin","BO":"Bolivia","BR":"Brazil","BW":"Botswana",
    "BY":"Belarus","BZ":"Belize","CA":"Canada","CD":"Congo (DRC)","CF":"Central African Republic","CG":"Congo",
    "CH":"Switzerland","CI":"Côte d'Ivoire","CL":"Chile","CM":"Cameroon","CN":"China","CO":"Colombia","CR":"Costa Rica",
    "CU":"Cuba","CV":"Cape Verde","CY":"Cyprus","CZ":"Czech Republic","DE":"Germany","DJ":"Djibouti","DK":"Denmark",
    "DO":"Dominican Republic","DZ":"Algeria","EC":"Ecuador","EE":"Estonia","EG":"Egypt","ES":"Spain","ET":"Ethiopia",
    "FI":"Finland","FR":"France","GA":"Gabon","GB":"United Kingdom","GE":"Georgia","GH":"Ghana","GM":"Gambia",
    "GN":"Guinea","GQ":"Equatorial Guinea","GR":"Greece","GT":"Guatemala","GW":"Guinea-Bissau","GY":"Guyana",
    "HN":"Honduras","HR":"Croatia","HU":"Hungary","ID":"Indonesia","IE":"Ireland","IL":"Israel","IN":"India",
    "IQ":"Iraq","IR":"Iran","IT":"Italy","JM":"Jamaica","JO":"Jordan","JP":"Japan","KE":"Kenya","KG":"Kyrgyzstan",
    "KH":"Cambodia","KM":"Comoros","KR":"South Korea","KW":"Kuwait","KZ":"Kazakhstan","LA":"Laos","LB":"Lebanon",
    "LK":"Sri Lanka","LR":"Liberia","LS":"Lesotho","LT":"Lithuania","LU":"Luxembourg","LV":"Latvia","LY":"Libya",
    "MA":"Morocco","MD":"Moldova","ME":"Montenegro","MG":"Madagascar","MK":"North Macedonia","ML":"Mali","MM":"Myanmar",
    "MN":"Mongolia","MR":"Mauritania","MT":"Malta","MU":"Mauritius","MW":"Malawi","MX":"Mexico","MY":"Malaysia",
    "MZ":"Mozambique","NA":"Namibia","NE":"Niger","NG":"Nigeria","NI":"Nicaragua","NL":"Netherlands","NO":"Norway",
    "NP":"Nepal","NZ":"New Zealand","OM":"Oman","PA":"Panama","PE":"Peru","PG":"Papua New Guinea","PH":"Philippines",
    "PK":"Pakistan","PL":"Poland","PT":"Portugal","PY":"Paraguay","QA":"Qatar","RO":"Romania","RS":"Serbia",
    "RU":"Russia","RW":"Rwanda","SA":"Saudi Arabia","SC":"Seychelles","SD":"Sudan","SE":"Sweden","SG":"Singapore",
    "SI":"Slovenia","SK":"Slovakia","SL":"Sierra Leone","SN":"Senegal","SO":"Somalia","SR":"Suriname","SV":"El Salvador",
    "SY":"Syria","SZ":"Eswatini","TD":"Chad","TG":"Togo","TH":"Thailand","TJ":"Tajikistan","TN":"Tunisia","TR":"Turkey",
    "TT":"Trinidad and Tobago","TW":"Taiwan","TZ":"Tanzania","UA":"Ukraine","UG":"Uganda","US":"United States",
    "UY":"Uruguay","UZ":"Uzbekistan","VE":"Venezuela","VN":"Vietnam","YE":"Yemen","ZA":"South Africa","ZM":"Zambia","ZW":"Zimbabwe"
}

def normalize_geo(text: str) -> str:
    s = normalize_spaces(text).lower().replace("’","'").replace("`","'").replace("´","'")
    s = strip_accents(s)
    return COUNTRY_ALIASES.get(s, s)

def normalize_site_id(value):
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    s = str(value).strip()
    if not s:
        return ""
    if re.fullmatch(r"\d+\.0+", s):
        s = s.split(".")[0]
    return s

def is_close_money(a, b, tol=0.5):
    if a is None or b is None:
        return False
    return abs(float(a) - float(b)) <= tol

def extract_number_after_keyword(line, keyword):
    m = re.search(rf"{keyword}\s*(?:,\s*\$)?\s*[:;=]?\s*\$?\s*([0-9]+(?:[.,][0-9]+)?)", line, flags=re.I)
    return safe_float(m.group(1)) if m else None

def parse_en_structured(line):
    if "country:" not in line.lower():
        return None
    site_m = re.search(r"\bSiteid\s+(\d+)\b", line, flags=re.I) or re.search(r"\bSiteid-(\d+)\b", line, flags=re.I) or re.search(r"\bWebsite\s*:\s*([^;,\n]+)", line, flags=re.I)
    geo_m = re.search(r"Country\s*:\s*([^;,\n]+)", line, flags=re.I)
    baseline_m = re.search(r"Baseline\s*:\s*([0-9]+(?:[.,][0-9]+)?)", line, flags=re.I)
    wager_m = re.search(r"Wager\s*:\s*([0-9]+(?:[.,][0-9]+)?)", line, flags=re.I)
    ftd_m = re.search(r"FTD\s*:\s*([0-9]+(?:[.,][0-9]+)?)", line, flags=re.I)
    cpa_m = re.search(r"CPA\s*(?:,\s*\$)?\s*[:;]\s*\$?\s*([0-9]+(?:[.,][0-9]+)?)", line, flags=re.I)
    spend_m = re.search(r"Spend\s*(?:,\s*\$)?\s*[:;]\s*\$?\s*([0-9]+(?:[.,][0-9]+)?)", line, flags=re.I)
    if not (geo_m and baseline_m and wager_m and cpa_m and ftd_m and spend_m):
        return None
    return {"status":"OK","parser_used":"en_structured","raw_line":line,"website":normalize_site_id(site_m.group(1)) if site_m else "",
            "geo":normalize_spaces(geo_m.group(1)),"baseline_type":"inclusive","baseline_value":safe_float(baseline_m.group(1)),
            "wager":safe_float(wager_m.group(1)),"rate":safe_float(cpa_m.group(1)),"manager_ftd":safe_int(ftd_m.group(1)),
            "manager_sum":safe_float(spend_m.group(1)),"comment":""}

def parse_short_cpa(line):
    if "cpa" not in line.lower() or "ftd" not in line.lower():
        return None
    working_line = line
    website = ""
    site_m = re.search(r"^\s*Siteid\s+(\d+)\s*;\s*", working_line, flags=re.I) or re.search(r"^\s*Siteid-(\d+)\s+", working_line, flags=re.I)
    if site_m:
        website = normalize_site_id(site_m.group(1))
        working_line = working_line[site_m.end():].strip()
    else:
        website_m = re.search(r"^\s*Website\s*:?\s*([^;\s,:]+)\s+", working_line, flags=re.I)
        if website_m:
            website = normalize_site_id(website_m.group(1))
            working_line = working_line[website_m.end():].strip()
    if re.search(r"\bmin\s*/\s*dep\b", working_line, flags=re.I):
        baseline_type = "positive"; baseline_value = 0.0
        geo = re.split(r"\bmin\s*/\s*dep\b", working_line, flags=re.I)[0].strip(" ,;")
    else:
        m_base = re.search(r"\bbase(?:line)?\s*:?\s*([0-9]+(?:[.,][0-9]+)?)", working_line, flags=re.I)
        if not m_base:
            return None
        baseline_type = "inclusive"; baseline_value = safe_float(m_base.group(1))
        geo = re.split(r"\bbase(?:line)?\s*:?\s*[0-9]+(?:[.,][0-9]+)?\b", working_line, flags=re.I)[0].strip(" ,;")
    wager = extract_number_after_keyword(working_line, "Wager")
    if wager is None:
        wager = 0.0
    cpa = extract_number_after_keyword(working_line, "CPA")
    ftd = extract_number_after_keyword(working_line, "FTD")
    spend = extract_number_after_keyword(working_line, "Spend")
    if not geo or cpa is None or ftd is None or spend is None:
        return None
    return {"status":"OK","parser_used":"short_cpa","raw_line":line,"website":website,"geo":normalize_spaces(geo),
            "baseline_type":baseline_type,"baseline_value":baseline_value,"wager":wager,"rate":cpa,
            "manager_ftd":safe_int(ftd),"manager_sum":safe_float(spend),"comment":""}

def parse_flexible(line):
    if not line:
        return None
    working_line = line; website = ""
    site_m = re.search(r"\bSiteid\s+(\d+)\b", working_line, flags=re.I) or re.search(r"\bSiteid-(\d+)\b", working_line, flags=re.I)
    if site_m:
        website = normalize_site_id(site_m.group(1)); working_line = working_line.replace(site_m.group(0), " ")
    else:
        web_m = re.search(r"\bWebsite\s*:?\s*(\d+)\b", working_line, flags=re.I)
        if web_m:
            website = normalize_site_id(web_m.group(1)); working_line = working_line.replace(web_m.group(0), " ")
    geo = re.split(r"\b(CPA|FTD|Spend|Base|Baseline|Wager|\$)\b", working_line, flags=re.I)[0].strip(" ,;")
    def find_num(keyword):
        m = re.search(rf"{keyword}\s*[:;]?\s*\$?\s*([0-9]+(?:[.,][0-9]+)?)", working_line, flags=re.I)
        return safe_float(m.group(1)) if m else None
    cpa = find_num("CPA"); ftd = find_num("FTD"); spend = find_num("Spend"); wager = find_num("Wager")
    m_base = re.search(r"(Base|Baseline)\s*[:;]?\s*([0-9]+(?:[.,][0-9]+)?)", working_line, flags=re.I)
    baseline_value = safe_float(m_base.group(2)) if m_base else 0.0
    if cpa is None:
        m_dollar = re.search(r"\$\s*([0-9]+(?:[.,][0-9]+)?)", working_line)
        if m_dollar:
            cpa = safe_float(m_dollar.group(1))
    if not geo or cpa is None or ftd is None or spend is None:
        return None
    return {"status":"OK","parser_used":"flexible","raw_line":line,"website":website,"geo":normalize_spaces(geo),
            "baseline_type":"inclusive","baseline_value":baseline_value,"wager":wager if wager is not None else 0.0,
            "rate":cpa,"manager_ftd":safe_int(ftd),"manager_sum":safe_float(spend),"comment":""}

def parse_geo_list_format(line):
    if "(" not in line:
        return None
    website = ""
    site_m = re.search(r"\bSite(?:id)?\s*[-: ]\s*(\d+)\b", line, flags=re.I)
    if site_m:
        website = normalize_site_id(site_m.group(1))
    pattern = re.compile(r"\b([A-Z]{2})\s*-\s*(?:\$)?\s*([0-9]+(?:[.,][0-9]+)?)\s*(?:\$)?\s*\(\s*([0-9]+)\s*\)", flags=re.I)
    matches = pattern.findall(line)
    if not matches:
        return None
    results = []
    for geo_code, spend, ftd in matches:
        spend_val = safe_float(spend); ftd_val = safe_int(ftd)
        if spend_val is None or ftd_val in (None, 0):
            continue
        results.append({"status":"OK","parser_used":"geo_list","raw_line":line,"website":website,
                        "geo":GEO_CODE_MAP.get(geo_code.upper(), geo_code.upper()),"baseline_type":"inclusive",
                        "baseline_value":0.0,"wager":0.0,"rate":spend_val/ftd_val,"manager_ftd":ftd_val,
                        "manager_sum":spend_val,"comment":""})
    return results if results else None

def parse_summary_line(line):
    line = normalize_spaces(line)
    if not line:
        return None
    low = line.lower()
    if low.startswith("period:") or low.startswith("total spend"):
        return {"status":"META","raw_line":line}
    for parser in [parse_en_structured, parse_short_cpa, parse_flexible, parse_geo_list_format]:
        try:
            result = parser(line)
            if result is not None:
                return result
        except Exception:
            pass
    return {"status":"ERROR","parser_used":"","raw_line":line,"website":"","geo":"","baseline_type":"",
            "baseline_value":None,"wager":None,"rate":None,"manager_ftd":None,"manager_sum":None,
            "comment":"Не удалось распознать строку сводки"}

def parse_summary_text(text):
    parsed, unrecognized = [], []
    for raw in text.splitlines():
        raw = normalize_spaces(raw)
        if not raw:
            continue
        row = parse_summary_line(raw)
        if row is None:
            continue
        if isinstance(row, list):
            for item in row:
                if item is None or item.get("status") == "META":
                    continue
                parsed.append(item)
                if item["status"] != "OK":
                    unrecognized.append(item)
        else:
            if row.get("status") == "META":
                continue
            parsed.append(row)
            if row["status"] != "OK":
                unrecognized.append(row)
    return parsed, unrecognized

def read_summary_from_second_sheet(xlsx_path):
    xls = pd.ExcelFile(xlsx_path)
    if len(xls.sheet_names) < 2:
        raise ValueError("Во входном файле нет второго листа со сводкой.")
    df = pd.read_excel(xlsx_path, sheet_name=xls.sheet_names[1], header=None, dtype=str)
    lines = []
    for _, row in df.iterrows():
        parts = []
        for val in row.tolist():
            if val is None:
                continue
            s = normalize_spaces(val)
            if s and s.lower() != "nan":
                parts.append(s)
        if parts:
            lines.append(" ".join(parts))
    summary_text = "\n".join(lines).strip()
    if not summary_text:
        raise ValueError("На втором листе не найдена сводка.")
    return summary_text

REQUIRED_COLS = {
    "player_id":["ID гравця","ID игрока","Player ID","ID"],
    "site_id":["ID сайта","ID сайту","Site ID","Website ID","Website","ID site","ID сайту"],
    "geo":["Країна","Страна","Country"],
    "deposit":["Сума депозитів","Сумма депозитов","Sum of all deposits","Deposit Sum","Deposits","Deposit"],
    "bets":["Сума ставок","Сумма ставок","Total bet amount","Bet Sum","Wager Sum","Wager","Bets"],
}

def detect_header_row(df_raw):
    best_idx, best_score = None, -1
    for i in range(min(30, len(df_raw))):
        row = [str(x).strip() for x in df_raw.iloc[i].tolist()]
        score = 0
        for variants in REQUIRED_COLS.values():
            for v in variants:
                if v in row:
                    score += 1
                    break
        if score > best_score:
            best_score, best_idx = score, i
    if best_idx is None or best_score < 3:
        raise ValueError("Не удалось найти строку заголовков в выгрузке.")
    return best_idx

def find_column_name(actual_columns, variants):
    for col in actual_columns:
        if str(col).strip() in variants:
            return col
    return None

def read_export_file(xlsx_path):
    df_raw = pd.read_excel(xlsx_path, sheet_name=0, header=None)
    header_row = detect_header_row(df_raw)
    df = pd.read_excel(xlsx_path, sheet_name=0, header=header_row).copy()
    player_col = find_column_name(df.columns, REQUIRED_COLS["player_id"])
    site_col = find_column_name(df.columns, REQUIRED_COLS["site_id"])
    geo_col = find_column_name(df.columns, REQUIRED_COLS["geo"])
    deposit_col = find_column_name(df.columns, REQUIRED_COLS["deposit"])
    bets_col = find_column_name(df.columns, REQUIRED_COLS["bets"])
    if not geo_col or not deposit_col or not bets_col:
        raise ValueError("Не найдены нужные колонки в выгрузке.")
    cleaned = pd.DataFrame()
    cleaned["player_id"] = df[player_col] if player_col else None
    cleaned["site_id"] = df[site_col].apply(normalize_site_id) if site_col else ""
    cleaned["geo"] = df[geo_col]
    cleaned["deposit"] = df[deposit_col].apply(safe_float)
    cleaned["bets"] = df[bets_col].apply(safe_float)
    cleaned = cleaned[cleaned["geo"].notna()].copy()
    cleaned["geo_raw"] = cleaned["geo"].astype(str).str.strip()
    cleaned["geo_norm"] = cleaned["geo_raw"].apply(normalize_geo)
    return cleaned, df

def is_valid_deposit(deposit, bets, baseline_type, baseline_value, wager):
    if deposit is None or bets is None or wager is None:
        return False
    if bets <= wager:
        return False
    if baseline_type == "positive":
        return deposit > 0
    if baseline_value is None:
        return False
    return deposit >= baseline_value

def build_comparison(parsed_rows, export_df):
    comparison_rows, eligible_rows = [], []
    for row in parsed_rows:
        if row["status"] != "OK":
            comparison_rows.append({"website":row.get("website",""),"geo":row.get("geo",""),"baseline_type":row.get("baseline_type",""),
                                    "baseline_value":row.get("baseline_value"),"wager":row.get("wager"),"rate":row.get("rate"),
                                    "manager_ftd":row.get("manager_ftd"),"actual_valid_ftd":None,"delta_ftd":None,
                                    "manager_sum":row.get("manager_sum"),"expected_sum_by_manager_ftd":None,"actual_sum":None,
                                    "sum_delta":None,"is_sum_ok":False,"is_ftd_ok":False,"status":"ERROR",
                                    "comment":row.get("comment",""),"parser_used":row.get("parser_used",""),"raw_line":row.get("raw_line","")})
            continue
        geo_norm = normalize_geo(row["geo"]); website = normalize_site_id(row.get("website",""))
        baseline_type = row["baseline_type"]; baseline_value = row["baseline_value"]; wager = row["wager"]
        rate = row["rate"]; manager_ftd = row["manager_ftd"]; manager_sum = row["manager_sum"]
        sub = export_df[export_df["geo_norm"] == geo_norm].copy()
        if website:
            sub = sub[sub["site_id"] == website].copy()
        if len(sub) > 0:
            sub["is_eligible"] = sub.apply(lambda x: is_valid_deposit(x["deposit"], x["bets"], baseline_type, baseline_value, wager), axis=1)
            eligible = sub[sub["is_eligible"]].copy()
        else:
            eligible = pd.DataFrame(columns=list(sub.columns)+["is_eligible"])
        actual_valid_ftd = int(len(eligible))
        delta_ftd = actual_valid_ftd - manager_ftd if manager_ftd is not None else None
        expected_sum_by_manager_ftd = float(rate) * float(manager_ftd) if rate is not None and manager_ftd is not None else None
        payable_ftd = min(actual_valid_ftd, manager_ftd) if manager_ftd is not None else actual_valid_ftd
        actual_sum = float(rate) * float(payable_ftd) if rate is not None else None
        sum_delta = float(manager_sum) - float(actual_sum) if manager_sum is not None and actual_sum is not None else None
        is_sum_ok = is_close_money(manager_sum, expected_sum_by_manager_ftd, tol=0.5)
        is_ftd_ok = actual_valid_ftd >= manager_ftd if manager_ftd is not None else False
        if is_ftd_ok and is_sum_ok:
            status = "OK"; comment = "По выгрузке депозитов достаточно, математика менеджера верная"
        elif (not is_ftd_ok) and is_sum_ok:
            status = "ERROR"; comment = "Выгрузка подтверждает меньше валидных депозитов, чем подал менеджер"
        elif is_ftd_ok and (not is_sum_ok):
            status = "ERROR"; comment = "По выгрузке депозитов достаточно, но математика менеджера неверная"
        else:
            status = "ERROR"; comment = "И депозитов не хватает, и математика менеджера неверная"
        comparison_rows.append({"website":website,"geo":row["geo"],"baseline_type":baseline_type,"baseline_value":baseline_value,
                                "wager":wager,"rate":rate,"manager_ftd":manager_ftd,"actual_valid_ftd":actual_valid_ftd,
                                "delta_ftd":delta_ftd,"manager_sum":manager_sum,"expected_sum_by_manager_ftd":expected_sum_by_manager_ftd,
                                "actual_sum":actual_sum,"sum_delta":sum_delta,"is_sum_ok":is_sum_ok,"is_ftd_ok":is_ftd_ok,
                                "status":status,"comment":comment,"parser_used":row["parser_used"],"raw_line":row["raw_line"]})
        if len(eligible) > 0:
            elig = eligible.copy()
            elig["summary_website"] = website; elig["summary_geo"] = row["geo"]; elig["summary_baseline_type"] = baseline_type
            elig["summary_baseline_value"] = baseline_value; elig["summary_wager"] = wager
            elig["summary_rate"] = rate; elig["summary_manager_ftd"] = manager_ftd
            eligible_rows.append(elig)
    return pd.DataFrame(comparison_rows), (pd.concat(eligible_rows, ignore_index=True) if eligible_rows else pd.DataFrame())

def autosize_worksheet(ws):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            val = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 60)

def format_sheet_header(ws):
    fill = PatternFill("solid", fgColor="1F4E78"); font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = fill; cell.font = font; cell.alignment = Alignment(horizontal="center", vertical="center")

def colorize_comparison(ws):
    ok_fill = PatternFill("solid", fgColor="E2F0D9")
    err_fill = PatternFill("solid", fgColor="FCE4D6")
    warn_fill = PatternFill("solid", fgColor="FFF2CC")
    headers = {cell.value: cell.column for cell in ws[1]}
    status_col = headers.get("status"); delta_col = headers.get("delta_ftd"); sum_delta_col = headers.get("sum_delta")
    for row in range(2, ws.max_row + 1):
        status = ws.cell(row=row, column=status_col).value if status_col else None
        delta_val = ws.cell(row=row, column=delta_col).value if delta_col else None
        sum_delta_val = ws.cell(row=row, column=sum_delta_col).value if sum_delta_col else None
        fill = ok_fill
        if status == "ERROR":
            fill = err_fill
        elif delta_val == 0:
            fill = warn_fill
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).fill = fill
        try:
            if sum_delta_val is not None and float(sum_delta_val) > 0:
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).fill = err_fill
        except Exception:
            pass

def write_excel_report(output_xlsx, parsed_preview_df, comparison_df, unrecognized_df, eligible_df, raw_export_df):
    with pd.ExcelWriter(output_xlsx, engine="openpyxl") as writer:
        comparison_df.to_excel(writer, sheet_name=COMPARISON_SHEET, index=False)
        parsed_preview_df.to_excel(writer, sheet_name=PREVIEW_SHEET, index=False)
        (pd.DataFrame([{"info":"Все строки распознаны"}]) if unrecognized_df is None or unrecognized_df.empty else unrecognized_df).to_excel(writer, sheet_name=UNRECOGNIZED_SHEET, index=False)
        (pd.DataFrame([{"info":"Нет подходящих игроков"}]) if eligible_df is None or eligible_df.empty else eligible_df).to_excel(writer, sheet_name=ELIGIBLE_SHEET, index=False)
        raw_export_df.to_excel(writer, sheet_name=RAW_SHEET, index=False)
        pd.DataFrame([
            {"rule":"Источник сводки","description":"Сводка читается со 2-го листа входного Excel-файла"},
            {"rule":"Валидный депозит","description":"deposit >= baseline (или > 0 для min/dep) и bets > wager"},
            {"rule":"FTD OK","description":"actual_valid_ftd >= manager_ftd"},
            {"rule":"SUM OK","description":"manager_sum ~= rate * manager_ftd (допуск 0.5)"},
            {"rule":"Website","description":"если Website/Siteid указан — фильтр по GEO + Website; если нет — только по GEO"},
            {"rule":"actual_sum","description":"sum to pay = min(actual_valid_ftd, manager_ftd) * rate"},
            {"rule":"sum_delta","description":"manager_sum - actual_sum"},
        ]).to_excel(writer, sheet_name=RULES_SHEET, index=False)
    wb = load_workbook(output_xlsx)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]; format_sheet_header(ws); autosize_worksheet(ws)
    if COMPARISON_SHEET in wb.sheetnames:
        colorize_comparison(wb[COMPARISON_SHEET])
    wb.save(output_xlsx)

def process_file(input_xlsx, output_xlsx):
    summary_text = read_summary_from_second_sheet(input_xlsx)
    parsed_rows, unrecognized_rows = parse_summary_text(summary_text)
    parsed_preview_df = pd.DataFrame(parsed_rows)
    export_df, raw_export_df = read_export_file(input_xlsx)
    comparison_df, eligible_df = build_comparison(parsed_rows, export_df)
    unrecognized_df = pd.DataFrame(unrecognized_rows)
    write_excel_report(output_xlsx, parsed_preview_df, comparison_df, unrecognized_df, eligible_df, raw_export_df)
    ok_cnt = int((comparison_df["status"] == "OK").sum()) if not comparison_df.empty else 0
    err_cnt = int((comparison_df["status"] == "ERROR").sum()) if not comparison_df.empty else 0
    return {"ok": ok_cnt, "error": err_cnt, "total": len(comparison_df), "output_xlsx": output_xlsx}
